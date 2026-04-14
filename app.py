import streamlit as st
import pandas as pd
import zipfile
import json
import io
import re
import datetime
from copy import copy
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from thefuzz import process, fuzz

st.set_page_config(page_title="Global AI Agent", layout="wide")

# --- DIZIONARIO MULTILINGUA ---
LANG = {
    "IT": {
        "title": "🤖 Global Data Automation Agent",
        "subtitle": "Data transfer with **Date Correction**, **Dynamic ZIP Search**, **Filters**, and **Safe Saving**.",
        "sidebar_title": "⚙️ Impostazioni",
        "select_data": "📂 Seleziona i dati da elaborare:",
        "upload_zip": "1. Carica lo ZIP dell'OCR",
        "upload_xls": "2. Carica il Master Pareto (XLSM)",
        "machines_loaded": "✅ {} Macchine in memoria",
        "file_found": " Elaborazione del file:\n{}",
        "err_no_json": "Nessun file 'output.json' trovato all'interno dello ZIP (nemmeno nelle sottocartelle).",
        "err_master": "Errore caricamento Master: {}",
        "warn_empty": "⚠️ Attenzione: Il file JSON selezionato è vuoto. L'OCR non ha estratto alcun dato.",
        "review_title": "📝 Revisione Dati",
        "filter_title": "🔍 Filtra i Dati (Stile Excel)",
        "filter_status": "Filtra per Stato:",
        "filter_tech": "Filtra per TechID:",
        "filter_err": "Filtra per Errore:",
        "btn_save": "🚀 Conferma, Filtra e Genera File",
        "err_no_data": "Non ci sono dati da salvare!",
        "err_no_sheet": "ATTENZIONE: Il foglio '{}' non esiste nel file!",
        "success_save": "Operazione completata! ✅ {} righe perfette aggiunte al Master. ❌ {} righe scartate.",
        "btn_download": "📥 Scarica File Aggiornato",
        "col_photo": "Foto Origine", "col_date": "Data", "col_tech": "TechID", 
        "col_event": "Evento", "col_code": "Codice Errore", "col_loss": "Perdita", 
        "col_notes": "Note Agente", "col_status": "Stato",
        "log_empty": "⚠️ Vuoto", "log_rebuilt": "Ricostruito", "log_fuzzy": "Fuzzy", 
        "log_unknown": "⚠️ Macchina ignota", "log_forced": "Forzato", "log_fix": "Code Fix",
        "stat_err": "❌ Errore", "stat_err_time": "❌ Errore (Tempo < 5)", "stat_err_invalid": "❌ Errore (Tempo non valido)", 
        "stat_ok": "✅ OK", "stat_ok_mod": "✅ OK (Corretto)"
    },
    "DE": {
        "title": "🤖 Global Data Automation Agent",
        "subtitle": "Data transfer with **Date Correction**, **Dynamic ZIP Search**, **Filters**, and **Safe Saving**.",
        "sidebar_title": "⚙️ Einstellungen",
        "select_data": "📂 Zu verarbeitende Daten wählen:",
        "upload_zip": "1. OCR ZIP-Datei hochladen",
        "upload_xls": "2. Pareto Master hochladen (XLSM)",
        "machines_loaded": "✅ {} Maschinen im Speicher",
        "file_found": " Verarbeitung der Datei:\n{}",
        "err_no_json": "Keine 'output.json'-Datei im ZIP gefunden (auch nicht in Unterordnern).",
        "err_master": "Fehler beim Laden des Masters: {}",
        "warn_empty": "⚠️ Achtung: Die gewählte JSON-Datei ist leer. Die OCR hat keine Daten extrahiert.",
        "review_title": "📝 Datenüberprüfung",
        "filter_title": "🔍 Daten filtern (Excel-Stil)",
        "filter_status": "Nach Status filtern:",
        "filter_tech": "Nach TechID filtern:",
        "filter_err": "Nach Fehlercode filtern:",
        "btn_save": "🚀 Bestätigen, Filtern & Datei erstellen",
        "err_no_data": "Keine Daten zum Speichern vorhanden!",
        "err_no_sheet": "ACHTUNG: Das Blatt '{}' existiert nicht in der Datei!",
        "success_save": "Vorgang abgeschlossen! ✅ {} perfekte Zeilen zum Master hinzugefügt. ❌ {} Zeilen verworfen.",
        "btn_download": "📥 Aktualisierte Datei herunterladen",
        "col_photo": "Originalfoto", "col_date": "Datum", "col_tech": "TechID", 
        "col_event": "Ereignis", "col_code": "Fehlercode", "col_loss": "Verlust (Min)", 
        "col_notes": "Agenten-Notiz", "col_status": "Status",
        "log_empty": "⚠️ Leer", "log_rebuilt": "Rekonstruiert", "log_fuzzy": "Fuzzy Match", 
        "log_unknown": "⚠️ Unbekannte Maschine", "log_forced": "Erzwungen", "log_fix": "Code Fix",
        "stat_err": "❌ Fehler", "stat_err_time": "❌ Fehler (Zeit < 5)", "stat_err_invalid": "❌ Fehler (Ungültige Zeit)", 
        "stat_ok": "✅ OK", "stat_ok_mod": "✅ OK (Korrigiert)"
    }
}

st.sidebar.title("🌍 Language / Lingua / Sprache")
lang_choice = st.sidebar.radio("", ["Italiano (IT)", "Deutsch (DE)"])
L = LANG["IT"] if "IT" in lang_choice else LANG["DE"]

st.title(L["title"])
st.markdown(L["subtitle"])

# --- INIZIALIZZAZIONE MEMORIA DI STATO ---
if "current_file" not in st.session_state:
    st.session_state.current_file = None
if "df_main" not in st.session_state:
    st.session_state.df_main = pd.DataFrame()

# --- CALCOLO DATA DI RIFERIMENTO E CALENDARIO SETTIMANA ---
oggi = datetime.date.today()
giorni_da_lunedi_scorso = oggi.weekday() + 7 
lunedi_scorso = oggi - datetime.timedelta(days=giorni_da_lunedi_scorso)

settimana_analisi = [lunedi_scorso + datetime.timedelta(days=i) for i in range(7)]
giorni_validi = {d.day: d for d in settimana_analisi}

def normalizza_data(data_grezza):
    data_str = str(data_grezza).strip()
    gg_estratto = None
    try:
        if "." in data_str:
            gg_estratto = int(data_str.split(".")[0])
        elif "/" in data_str:
            gg_estratto = int(data_str.split("/")[0])
        elif data_str.isdigit():
            gg_estratto = int(data_str[:2])
    except:
        pass 
        
    if gg_estratto in giorni_validi:
        return giorni_validi[gg_estratto].strftime("%d/%m/%Y")
    return lunedi_scorso.strftime("%d/%m/%Y")

def pulisci_minuti(valore_grezzo):
    if pd.isna(valore_grezzo) or str(valore_grezzo).strip() == "":
        return 0.0
    val = str(valore_grezzo).replace(',', '.')
    val_pulito = re.sub(r'[^0-9.]', '', val)
    try:
        return float(val_pulito) if val_pulito else 0.0
    except ValueError:
        return 0.0

def suggerisci_correzione(valore, lista_validazione, soglia=75):
    if pd.isna(valore) or str(valore).strip() == "":
        return None
    valore_str = str(valore).strip().upper()
    miglior_match, punteggio = process.extractOne(valore_str, lista_validazione, scorer=fuzz.ratio)
    if punteggio >= soglia:
        return miglior_match
    return None

def formatta_tech_id(valore_grezzo, lista_validazione):
    if pd.isna(valore_grezzo) or str(valore_grezzo).strip() == "":
        return "", L["log_empty"]
        
    valore = str(valore_grezzo).upper().replace(" ", "")
    valore = valore.replace("/\\", "1").replace("\\/", "1")
    valore = re.sub(r'[\\/|]+', '1', valore)
    
    valore_pulito = re.sub(r'[^A-Z0-9]', '', valore)
    if not valore_pulito:
        return "", L["log_empty"]

    if valore_pulito in lista_validazione:
        return valore_pulito, ""
    
    lettere = "".join([c for c in valore_pulito if c.isalpha()])
    numeri = "".join([c for c in valore_pulito if c.isdigit()])
    
    if lettere and numeri:
        id_opt1 = lettere[:2] + numeri[-3:].zfill(3)
        id_opt2 = lettere[:3] + numeri[-2:].zfill(2)
        
        if id_opt1 in lista_validazione:
            return id_opt1, L["log_rebuilt"]
        if id_opt2 in lista_validazione:
            return id_opt2, L["log_rebuilt"]
            
    if len(lista_validazione) > 0:
        miglior_match, punteggio = process.extractOne(valore_pulito, lista_validazione, scorer=fuzz.ratio)
        if punteggio >= 65:
            return miglior_match, f"{L['log_fuzzy']} {punteggio}%"
            
    return "", L["log_unknown"]

# --- INTERFACCIA E CARICAMENTO FILE ---
col1, col2 = st.columns(2)
with col1:
    zip_file = st.file_uploader(L["upload_zip"], type="zip")
with col2:
    master_file = st.file_uploader(L["upload_xls"], type=["xlsm", "xlsx"])

percorso_selezionato = None

if zip_file:
    try:
        with zipfile.ZipFile(zip_file, 'r') as z:
            elenco_file = z.namelist()
            possibili_json = [f for f in elenco_file if f.endswith('output.json')]
            
            if possibili_json:
                st.sidebar.markdown("---")
                st.sidebar.title(L["sidebar_title"])
                percorso_selezionato = st.sidebar.selectbox(L["select_data"], options=possibili_json)
            else:
                st.sidebar.error(L["err_no_json"])
    except Exception as e:
        st.sidebar.error(f"Error reading ZIP / Errore lettura ZIP: {e}")

if zip_file and master_file and percorso_selezionato:
    try:
        df_equip = pd.read_excel(master_file, sheet_name='EquipmentLIST')
        df_codes = pd.read_excel(master_file, sheet_name='CODE-ERR')
        
        valid_tech_ids = df_equip['TechID'].astype(str).str.strip().str.upper().tolist()
        valid_error_codes = df_codes['Code'].astype(str).str.strip().str.upper().tolist()
        st.sidebar.success(L["machines_loaded"].format(len(valid_tech_ids)))
        st.sidebar.info(L["file_found"].format(percorso_selezionato))
    except Exception as e:
        st.error(L["err_master"].format(e))
        st.stop()

    # MOTORE DI VALUTAZIONE DINAMICA DELLO STATO
    def evaluate_status(row):
        is_modded = False
        try:
            curr_loss = float(row[L["col_loss"]]) if pd.notna(row[L["col_loss"]]) and str(row[L["col_loss"]]).strip() != "" else 0.0
            orig_loss = float(row["_Orig_Loss"]) if pd.notna(row["_Orig_Loss"]) and str(row["_Orig_Loss"]).strip() != "" else 0.0
            
            if str(row[L["col_tech"]]) != str(row["_Orig_Tech"]) or str(row[L["col_code"]]) != str(row["_Orig_Code"]) or curr_loss != orig_loss:
                is_modded = True
        except:
            pass

        t_id = str(row[L["col_tech"]]).strip().upper()
        c_err = str(row[L["col_code"]]).strip().upper()

        if pd.isna(row[L["col_tech"]]) or pd.isna(row[L["col_code"]]) or t_id == "❓" or t_id == "":
            return L["stat_err"]
        
        try:
            if float(row[L["col_loss"]]) < 5:
                return L["stat_err_time"]
        except:
            return L["stat_err_invalid"]
            
        if t_id in valid_tech_ids and c_err in valid_error_codes:
            if is_modded:
                return L["stat_ok_mod"]
            return L["stat_ok"]
            
        return L["stat_err"]

    # CARICAMENTO INIZIALE (Avviene solo quando cambi il percorso del file selezionato)
    if percorso_selezionato != st.session_state.current_file:
        with zipfile.ZipFile(zip_file, 'r') as z:
            with z.open(percorso_selezionato) as f:
                dati_ocr = json.load(f)
            
            rows = []
            for entry in dati_ocr:
                tech_id_raw = entry.get('tech_id', '').strip()
                img_name = entry.get('image_name_oryginal', 'Unknown')
                
                tech_id_final, nota_tech = formatta_tech_id(tech_id_raw, valid_tech_ids)
                display_tech_id = tech_id_final if tech_id_final else "❓"

                for r in entry.get('rows', []):
                    codice_raw = str(r.get('code', '')).strip().upper()
                    codice_final = codice_raw
                    nota_code = ""
                    
                    if codice_raw in ["C06", "E01"]:
                        if "C06/E01" in valid_error_codes:
                            codice_final = "C06/E01"
                        elif "E01/C06" in valid_error_codes:
                            codice_final = "E01/C06"
                        else:
                            codice_final = "C06/E01"
                        nota_code = f"{L['log_forced']} C06/E01"
                    elif codice_raw not in valid_error_codes:
                        suggerimento_c = suggerisci_correzione(codice_raw, valid_error_codes)
                        if suggerimento_c:
                            codice_final = suggerimento_c
                            nota_code = L["log_fix"]
                    
                    data_pulita = normalizza_data(r.get('date', ''))
                    minuti_puliti = pulisci_minuti(r.get('minutes_lost', 0))
                    nota_completa = " | ".join(filter(None, [nota_tech, nota_code]))
                    
                    rows.append({
                        L["col_photo"]: img_name,
                        L["col_date"]: data_pulita,
                        L["col_tech"]: display_tech_id,
                        L["col_event"]: str(r.get('event', '')).strip(),
                        L["col_code"]: codice_final,
                        L["col_loss"]: minuti_puliti,
                        L["col_notes"]: nota_completa
                    })
            
            colonne_base = [L["col_photo"], L["col_date"], L["col_tech"], L["col_event"], L["col_code"], L["col_loss"], L["col_notes"]]
            df_base = pd.DataFrame(rows, columns=colonne_base)
            
            if not df_base.empty:
                df_base["_Orig_Tech"] = df_base[L["col_tech"]]
                df_base["_Orig_Code"] = df_base[L["col_code"]]
                df_base["_Orig_Loss"] = df_base[L["col_loss"]]
                df_base[L["col_status"]] = df_base.apply(evaluate_status, axis=1)
            else:
                df_base[L["col_status"]] = pd.Series(dtype=str)
                st.warning(L["warn_empty"])

            st.session_state.df_main = df_base
            st.session_state.current_file = percorso_selezionato

    # --- FUNZIONE FRAGMENT (ISOLA IL REFRESH DELLA TABELLA) ---
    @st.fragment
    def render_editor_section():
        df_main = st.session_state.df_main

        st.subheader(L["review_title"])
        
        with st.expander(L["filter_title"], expanded=True):
            col_f1, col_f2, col_f3 = st.columns(3)
            
            opzioni_stato = df_main[L["col_status"]].unique().tolist() if not df_main.empty else []
            opzioni_tech = df_main[L["col_tech"]].unique().tolist() if not df_main.empty else []
            opzioni_err = df_main[L["col_code"]].unique().tolist() if not df_main.empty else []
            
            filtro_stato = col_f1.multiselect(L["filter_status"], options=opzioni_stato, default=[])
            filtro_techid = col_f2.multiselect(L["filter_tech"], options=opzioni_tech, default=[])
            filtro_errore = col_f3.multiselect(L["filter_err"], options=opzioni_err, default=[])
            
        df_filtrato = df_main.copy()
        if filtro_stato:
            df_filtrato = df_filtrato[df_filtrato[L["col_status"]].isin(filtro_stato)]
        if filtro_techid:
            df_filtrato = df_filtrato[df_filtrato[L["col_tech"]].isin(filtro_techid)]
        if filtro_errore:
            df_filtrato = df_filtrato[df_filtrato[L["col_code"]].isin(filtro_errore)]
        
        df_edited_filtered = st.data_editor(
            df_filtrato, 
            use_container_width=True, 
            num_rows="dynamic",
            column_config={
                L["col_photo"]: st.column_config.TextColumn("File", disabled=True),
                L["col_status"]: st.column_config.TextColumn("Status", disabled=True),
                L["col_notes"]: st.column_config.TextColumn("IA Log", disabled=True),
                L["col_loss"]: st.column_config.NumberColumn("Min.", format="%.2f"),
                "_Orig_Tech": None,  
                "_Orig_Code": None,
                "_Orig_Loss": None
            }
        )

        # --- RICUCITURA DATI E RICALCOLO DINAMICO ---
        if not df_main.empty and not df_edited_filtered.empty:
            vecchio_status = st.session_state.df_main[L["col_status"]].copy()
            
            st.session_state.df_main.update(df_edited_filtered)
            st.session_state.df_main[L["col_status"]] = st.session_state.df_main.apply(evaluate_status, axis=1)

            if not st.session_state.df_main[L["col_status"]].equals(vecchio_status):
                st.rerun(scope="fragment")
                
    # Richiamo il Fragment
    if "df_main" in st.session_state and not st.session_state.df_main.empty:
        render_editor_section()

    # --- SALVATAGGIO DEFINITIVO ---
    if st.button(L["btn_save"]):
        if st.session_state.df_main.empty:
            st.error(L["err_no_data"])
            st.stop()
            
        try:
            output_buffer = io.BytesIO(master_file.getvalue())
            
            # --- FIX PER ERRORE EXCEL "externalLink1.xml" ---
            # keep_links=False impedisce a openpyxl di corrompere i link esterni presenti nel file
            book = load_workbook(output_buffer, keep_vba=True, keep_links=False)
            
            nome_foglio = 'ParetoDATA'
            if nome_foglio not in book.sheetnames:
                st.error(L["err_no_sheet"].format(nome_foglio))
                st.stop()
                
            ws = book[nome_foglio]
            
            target_table = None
            for tbl in ws.tables.values():
                if "pareto" in tbl.name.lower() or "pareto" in tbl.displayName.lower():
                    target_table = tbl
                    break
            if not target_table and len(ws.tables) > 0:
                target_table = list(ws.tables.values())[0]

            next_row = 1
            for row in range(ws.max_row, 0, -1):
                if ws.cell(row=row, column=12).value is not None or ws.cell(row=row, column=13).value is not None:
                    next_row = row + 1
                    break
            
            righe_salvate = 0
            righe_scartate = 0
            colonne_dati = [12, 13, 14, 15, 16] 
            
            for r in st.session_state.df_main.to_dict('records'):
                t_id = str(r.get(L['col_tech'], '')).strip().upper()
                c_err = str(r.get(L['col_code'], '')).strip().upper()
                
                try:
                    perdita_val = float(r.get(L['col_loss'], 0.0))
                except:
                    perdita_val = 0.0
                
                if t_id in valid_tech_ids and c_err in valid_error_codes and perdita_val >= 5:
                    # Inserimento della data come stringa per evitare sfasamenti, ma in futuro se i colleghi notano i "triangolini verdi"
                    # si può usare datetime.strptime(r.get(L['col_date'], ''), "%d/%m/%Y")
                    ws.cell(row=next_row, column=12).value = r.get(L['col_date'], '')         
                    ws.cell(row=next_row, column=13).value = t_id                      
                    try:
                        ws.cell(row=next_row, column=14).value = int(r.get(L['col_event'], 0))
                    except:
                        ws.cell(row=next_row, column=14).value = r.get(L['col_event'], '')    
                    ws.cell(row=next_row, column=15).value = c_err                     
                    ws.cell(row=next_row, column=16).value = perdita_val               
                    
                    for col in range(1, ws.max_column + 1):
                        prev_cell = ws.cell(row=next_row - 1, column=col)
                        new_cell = ws.cell(row=next_row, column=col)
                        
                        if prev_cell.has_style:
                            new_cell.font = copy(prev_cell.font)
                            new_cell.border = copy(prev_cell.border)
                            new_cell.fill = copy(prev_cell.fill)
                            new_cell.number_format = copy(prev_cell.number_format)
                            new_cell.alignment = copy(prev_cell.alignment)
                        
                        if col not in colonne_dati:
                            if isinstance(prev_cell.value, str) and prev_cell.value.startswith('='):
                                new_cell.value = Translator(prev_cell.value, origin=prev_cell.coordinate).translate_formula(new_cell.coordinate)
                    
                    next_row += 1
                    righe_salvate += 1
                else:
                    righe_scartate += 1
                    
            if target_table:
                ref_attuale = target_table.ref 
                if ":" in ref_attuale:
                    start_cell, end_cell = ref_attuale.split(":")
                    end_col_letter = "".join([c for c in end_cell if c.isalpha()])
                    nuova_ref = f"{start_cell}:{end_col_letter}{next_row - 1}"
                    
                    target_table.ref = nuova_ref
                    if target_table.autoFilter:
                        target_table.autoFilter.ref = nuova_ref
                
                target_table.sortState = None 

            for sheet in book.worksheets:
                try:
                    if sheet.auto_filter:
                        sheet.auto_filter.ref = None
                except:
                    pass

            final_output = io.BytesIO()
            book.save(final_output)
            
            st.success(L["success_save"].format(righe_salvate, righe_scartate))
            st.download_button(
                label=L["btn_download"],
                data=final_output.getvalue(),
                file_name="Pareto_Update_Global.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12"
            )
        except Exception as e:
            st.error(f"Error: {e}")
